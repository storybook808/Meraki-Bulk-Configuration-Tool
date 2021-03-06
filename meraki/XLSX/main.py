import csv
import merakiapi
import openpyxl

class Device:
    def __init__(self, row):
        self.serial_number = row[0]
        self.hostname = row[1]
        self.ip = row[2]
        self.vlan = row[3]
        self.subnet_mask = row[4]
        self.primary_dns = row[5]
        self.secondary_dns = row[6]
        self.tags = row[7]

class SwitchPort:
    def __init__(self, row):
        self.hostname = row[0]
        self.serial_number = row[1]
        self.port_number = row[2]
        self.name = row[3]
        self.tags = row[4]

        if row[5].lower() == "true":
            self.enabled = True
        else:
            self.enabled = False

        if row[6].lower() == "true":
            self.rstp = True
        else:
            self.rstp = False

        self.stp_guard = row[7]

        if row[8].lower() == "true":
            self.poe = True
        else:
            self.poe = False

        self.type = row[9]
        self.vlan = row[10]
        self.voice_vlan = row[11]
        self.allowed_vlan = row[12]

# def main():
#     # Pull the configurations.
#     configurations = {}
#     file_1 = open("Device Base Configuration.csv")
#     csv_1 = csv.reader(file_1)
#     for row in csv_1:
#         configurations[row[0]] = Device(row)
#
#     # API key.
#     api_key = ""
#
#     # Get the organization name.
#     print("Organization Name:")
#     org_name = raw_input()
#
#     # Pull the organizations associated to the provided API key.
#     orgs = merakiapi.myorgaccess(api_key, True)
#
#     # Look for the organization that we want to configure.
#     org_id = ""
#     for org in orgs:
#         if org_name in org["name"]:
#             org_id = org["id"]
#
#     if org_id == "":
#         print("Organization not found.")
#         return
#
#     # Pull the networks associated with the organization.
#     networks = merakiapi.getnetworklist(api_key, org_id, True)
#
#     # Pull the devices from all of the networks.
#     devices = []
#     for network in networks:
#         devices += merakiapi.getnetworkdevices(api_key, network["id"], True)
#
#     # Apply configuration to the devices and push them to Meraki.
#     for device in devices:
#         device["name"] = configurations[device["serial"]].hostname
#         merakiapi.updatedevice(api_key, device["networkId"], device["serial"], device["name"], "", "", "", "", True)
#
#     return

def main():
    # read from xlsx file
    configurations = {}
    from openpyxl import load_workbook
    wb = load_workbook(filename='template.xlsx')
    first_sheet = wb.get_sheet_names()[0]
    ws = wb[first_sheet]

    my_row = []
    for i in range(1, ws.max_row - 2):
        for j in range(1, ws.max_column - 1):
            my_row.append(ws.cell(row=i+1, column=j).value)

        #configurations[ws.cell(row=i + 1, column=2).value + str(ws.cell(row=i + 1, column=3).value)] = SwitchPort(my_row)  # dictionary
        print(ws.cell(row=i + 1, column=2).value + str(ws.cell(row=i + 1, column=3).value))
        print(my_row)

    print (configurations)

    # API key.
    api_key = "8b43aaa7b92b6d3ad06234e6f581077620d3e512"

    # Get the organization name.
    print("Organization Name:")
    org_name = input()

    # Pull the organizations associated to the provided API key.
    orgs = merakiapi.myorgaccess(api_key, True)

    # Look for the organization that we want to configure.
    org_id = ""
    for org in orgs:
        if org_name in org["name"]:
            org_id = org["id"]

    if org_id == "":
        print("Organization not found.")
        return

    # Pull the networks associated with the organization.
    networks = merakiapi.getnetworklist(api_key, org_id, True)

    # Pull the devices from all of the networks.
    devices = []
    for network in networks:
        devices += merakiapi.getnetworkdevices(api_key, network["id"], True)

    # print devices

    switch_ports = []
    for device in devices:
        current_switch_ports = []
        if device["model"].startswith("MS"):
            # current_switch_port = merakiapi.getswitchports(api_key, device["serial"])
            # current_switch_port["serial"] = device["serial"]
            current_switch_ports = merakiapi.getswitchports(api_key, device["serial"], False)

        # Label all current switch ports with the serial number of the parent switch.
        for switch_port in current_switch_ports:
            switch_port["serial"] = device["serial"]

        # Append the switch ports for the current switch to the master list.
        switch_ports += current_switch_ports

    print (switch_ports)

    # Apply configuration to the devices and push them to Meraki.
    for switch_port in switch_ports:
        try:
            switch_port["name"] = configurations[switch_port["serial"]+str(switch_port["number"])].name
        except:
            continue
        switch_port["tags"]  = configurations[switch_port["serial"]+str(switch_port["number"])].tags
        switch_port["enabled"] = configurations[switch_port["serial"]+str(switch_port["number"])].enabled
        switch_port["rstpEnabled"] = configurations[switch_port["serial"]+str(switch_port["number"])].rstp
        switch_port["stpGuard"] = configurations[switch_port["serial"]+str(switch_port["number"])].stp_guard
        switch_port["poeEnabled"] = configurations[switch_port["serial"]+str(switch_port["number"])].poe
        switch_port["type"] = configurations[switch_port["serial"]+str(switch_port["number"])].type
        switch_port["vlan"] = configurations[switch_port["serial"]+str(switch_port["number"])].vlan
        switch_port["voiceVlan"] = configurations[switch_port["serial"]+str(switch_port["number"])].voice_vlan
        switch_port["allowedVlans"] = configurations[switch_port["serial"]+str(switch_port["number"])].allowed_vlan

        # print switch_port["enabled"]


        merakiapi.updateswitchport(api_key, switch_port["serial"], switch_port["number"], switch_port["name"],
                                   switch_port["tags"], switch_port["enabled"], switch_port["type"],
                                   switch_port["vlan"], switch_port["voiceVlan"], switch_port["allowedVlans"],
                                   switch_port["poeEnabled"], "", switch_port["rstpEnabled"], switch_port["stpGuard"],
                                   "")



if __name__ == "__main__":
    main()