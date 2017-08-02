from flask import Blueprint
import merakiapi ,time
import os, shutil
from flask import Flask, stream_with_context, request, Response, flash, send_file
from time import sleep
from app import app
import xlrd

configure_blueprint = Blueprint('configure', __name__, template_folder='templates')

@app.route('/step3')
def progress_bar():
    global progress_bar
    return send_file('templates/step3.html')


# This function configures the meraki page
@configure_blueprint.route('/progress')
def configure():

    global progress_percent
    global org_name


# This function configures the meraki page

    class Device:
        def __init__(self, row):
            self.hostname = row[0]
            self.serial_number = row[1]
            self.ip = row[2]
            self.vlan = row[3]
            self.subnet_mask = row[4]
            self.primary_dns = row[5]
            self.secondary_dns = row[6]
            self.tags = row[7]

    class SwitchPort:
        # Constructor for switch port
        def __init__(self, row):
            self.hostname = row[0]
            self.serial_number = row[1]
            self.port_number = row[2]
            self.name = row[3]
            self.tags = row[4]

            if row[5] == "TRUE":
                self.enabled = True
            else:
                self.enabled = False

            if row[6] == "TRUE":
                self.rstp = True
            else:
                self.rstp = False

            self.stp_guard = row[7]

            if row[8] == "TRUE":
                self.poe = True
            else:
                self.poe = False

            self.type = row[9]
            self.vlan = row[10]
            self.voice_vlan = row[11]
            self.allowed_vlan = row[12]


    def get_api_key(api_path):
        workbook = xlrd.open_workbook(api_path)
        ws = workbook.sheet_by_index(0)
        cell = ws.cell_value(1, 0)
        print(cell)
        print(type(cell))
        return cell




    # Time Function
    # Purpose: Calculate the time to append to file name to better
    # debug Meraki configurations
    # Param: None
    # Output: Date_Time

    def time():
        import time

        date = time.strftime("%d.%m.%Y")
        time = time.strftime("%H.%M.%S")

        return date + "_" + time

    # sessionID Function
    # Purpose: return a unique ID tag to see what configuration of the Meraki
    # switch is related to
    # Param: None
    # Output: 10 character string of uppercase chars
    # def sessionID():
    #   import string
    #    import random

    #    chars = string.ascii_uppercase
    #    size = 4

    #    return ''.join(random.choice(chars) for _ in range(size))

    # file_rename Function
    # Purpose: renames files using the above time and sessionID functions
    # Param: None
    # Output: renamed file put into archive

    def file_rename():

        # assign a path based to the temp folder
        path = os.path.abspath(os.path.join('app', 'temp'))
        # get a list of all files in temp
        # this should only be a single file
        current_file = os.listdir(path)
        # id = sessionID()
        # print("file_RE_NAME")
        # print(current_file)
        print(os.path.abspath(os.path.join('temp', current_file[0])))
        # get the absolute path to the singleton file
        rename_src_path = os.path.abspath(os.path.join("app", "temp", current_file[0]))
        # get the absolute path to the destination folder, archive
        # name the new file based on the time and a unique ID
        rename_dst_path = os.path.abspath(
            os.path.join('app', 'archive',
                         current_file[0].replace(".xlsx", "") + "_" + time() + ".xlsx"))

        copy_dst_path = os.path.abspath(
            os.path.join('app', 'temp', current_file[0]))
        shutil.copy(rename_src_path, rename_dst_path)
        os.replace(rename_src_path, copy_dst_path)


       # if len(archive_folder) > 2:
        #    while len(archive_folder) > 2:
        #        os.remove(os.path.join(archive_path, archive_folder[0]))

    # archive_limit() Function
    # Purpose: limits the amount of files allowed in the archive folder
    # preventing the archive folder becoming unnecessarily large
    # Param: None
    # Output: None, but files will be removed
    def archive_limit():

        # set a counter varible for use in the while loop
        counter = 0

        # declare a path to the archive folder
        archive_path = os.path.abspath(os.path.join('app', 'archive'))

        # using the archive_path list the elements of the archive folder using os.listdir(path)
        archive_folder = os.listdir(archive_path)

        # initialize a list to hold the files that will be sorted/delted
        arch_folder_path_list = []

        # populate the list
        for i in range(len(archive_folder)):
            arch_folder_path_list.append(os.path.join(archive_path, archive_folder[i]))
        # using a bubble sort, sort the list based on the date last modified
        for j in range(len(arch_folder_path_list)):
            for k in range(len(arch_folder_path_list)-1, j, -1):
                if os.path.getmtime(arch_folder_path_list[k]) < os.path.getmtime(arch_folder_path_list[k-1]):
                    arch_folder_path_list[k], arch_folder_path_list[k-1] = arch_folder_path_list[k-1], arch_folder_path_list[k]

        # if the size of the list is over 20 delete the oldest files until the file size is 20
        if len(arch_folder_path_list) >= 20:
            while len(arch_folder_path_list) > 19:
                os.remove(arch_folder_path_list[counter])
                print(arch_folder_path_list[counter])
                counter += 1
                if counter == 19:
                    break

    ### Code to configure the base configuration for switches
    # def main():
    #     # Pull the configurations.
    #     configurations = {}
    #     file_1 = open("Device Base Configuration.csv")
    #     csv_1 = csv.reader(file_1)
    #     for row in csv_1:
    #         configurations[row[0]] = Device(row)
    #
    #     # API key.
    #     api_key = ""
    #
    #     # Get the organization name.
    #     print("Organization Name:")
    #     org_name = raw_input()
    #
    #     # Pull the organizations associated to the provided API key.
    #     orgs = merakiapi.myorgaccess(api_key, True)
    #
    #     # Look for the organization that we want to configure.
    #     org_id = ""
    #     for org in orgs:
    #         if org_name in org["name"]:
    #             org_id = org["id"]
    #
    #     if org_id == "":
    #         print("Organization not found.")
    #         return
    #
    #     # Pull the networks associated with the organization.
    #     networks = merakiapi.getnetworklist(api_key, org_id, True)
    #
    #     # Pull the devices from all of the networks.
    #     devices = []
    #     for network in networks:
    #         devices += merakiapi.getnetworkdevices(api_key, network["id"], True)
    #
    #     # Apply configuration to the devices and push them to Meraki.
    #     for device in devices:
    #         device["name"] = configurations[device["serial"]].hostname
    #         merakiapi.updatedevice(api_key, device["networkId"], device["serial"], device["name"], "", "", "", "", True)
    #
    #     return

    ### Find file path to pull configurations ###
    path = os.path.abspath(os.path.join('app', 'temp'))
    # path = os.path.join(initial_path, 'temp')
    current_file = os.listdir(path)
    # print(initial_path)
    # print(path)
    # print(current_file)
    print(current_file[0])
    print(os.path.join(path, current_file[0]))
    temp_path = os.path.join(path, current_file[0])  # path of configuration file

    ### Pull the configurations. ###
    configurations = {}
    from openpyxl import load_workbook
    wb = load_workbook(filename=temp_path)
    port_config_sheet = wb.get_sheet_names()[1]
    print(port_config_sheet)
    ws = wb[port_config_sheet]

    ### valid row count ###
    valid_row = 0  # will be incremented to the number of valid rows
    threshold = 2  # how many consecutive blank rows are allow before program stops scanning excel content
    real_row = 0  # track the actual row in excel file, including blank rows
    th_count = threshold
    while (th_count != 0):
        real_row += 1
        if (ws.cell(row=real_row, column=2).value != None and ws.cell(row=real_row, column=3).value != None):
            # print(ws.cell(row = valid_row + 1, column=2).value)
            valid_row += 1
            th_count = threshold  # if blank row is not consecutive to the number of threshold, reset th_count
        elif (ws.cell(row=real_row, column=2).value == None and ws.cell(row=real_row, column=3).value != None):
            print("row", real_row, "has incomplete entry")
        elif (ws.cell(row=real_row, column=2).value != None and ws.cell(row=real_row, column=3).value == None):
            print("row", real_row, "has incomplete entry")
        else:
            th_count -= 1

    progress_total = valid_row - 1  # save total number of excel entries to track for progress while compiling


    # create dictionary for switch port
    my_row = []
    i = 0
    while valid_row != 0:
        i += 1
        if ws.cell(row=i, column=2).value == None and ws.cell(row=i, column=3).value != None:
            pass
        elif ws.cell(row=i, column=2).value != None and ws.cell(row=i, column=3).value == None:
            pass
        elif ws.cell(row=i, column=2).value == None and ws.cell(row=i, column=3).value == None:
            pass
        else:
            for j in range(1, 14):  # max column number
                my_row.append(ws.cell(row=i, column=j).value)
            configurations[ws.cell(row=i, column=2).value, str(ws.cell(row=i, column=3).value)] = SwitchPort(
                my_row)  # dictionary
            # print(ws.cell(row = i, column = 2).value , str(ws.cell(row = i, column=3).value))
            # print(my_row)
            my_row = []  # reset
            valid_row -= 1

    # print(configurations)

    # API key.


    api_key = get_api_key(temp_path)



    ### Pull the organizations associated to the provided API key.
    orgs = merakiapi.myorgaccess(api_key, True)
    print(orgs)

    org_name = "World Wide"

    for org in orgs:
        if org_name in org["name"]:
            org_id = org["id"]

    if org_id == "":
        print("Orginization not Found")
    ### Pull the networks associated with the organization. ###
    networks = []
    for org in orgs:
        networks += merakiapi.getnetworklist(api_key, org["id"], True)

    ### Pull the devices from all of the networks. ###
    devices = []
    for network in networks:
        devices += merakiapi.getnetworkdevices(api_key, network["id"], True)

    switch_ports = []
    for device in devices:
        current_switch_ports = []
        if device["model"].startswith("MS"):
            # current_switch_port = merakiapi.getswitchports(api_key, device["serial"])
            # current_switch_port["serial"] = device["serial"]
            current_switch_ports = merakiapi.getswitchports(api_key, device["serial"], True)

        # Label all current switch ports with the serial number of the parent switch.
        for switch_port in current_switch_ports:
            switch_port["serial"] = device["serial"]

        # Append the switch ports for the current switch to the master list.
        switch_ports += current_switch_ports

    print(switch_ports)

    ### Apply configuration to the devices and push them to Meraki. ###

    ### Yield progress bar status to site ###
    def generate():
        progress_count = 0
        for switch_port in switch_ports:
            try:
                switch_port["name"] = configurations[switch_port["serial"], str(switch_port["number"])].name
            except:
                continue
            switch_port["tags"] = configurations[switch_port["serial"], str(switch_port["number"])].tags
            switch_port["enabled"] = configurations[switch_port["serial"], str(switch_port["number"])].enabled
            switch_port["rstpEnabled"] = configurations[switch_port["serial"], str(switch_port["number"])].rstp
            switch_port["stpGuard"] = configurations[switch_port["serial"], str(switch_port["number"])].stp_guard
            switch_port["poeEnabled"] = configurations[switch_port["serial"], str(switch_port["number"])].poe
            switch_port["type"] = configurations[switch_port["serial"], str(switch_port["number"])].type
            switch_port["vlan"] = configurations[switch_port["serial"], str(switch_port["number"])].vlan
            switch_port["voiceVlan"] = configurations[switch_port["serial"], str(switch_port["number"])].voice_vlan
            switch_port["allowedVlans"] = configurations[switch_port["serial"], str(switch_port["number"])].allowed_vlan

            # print (switch_port["enabled"])


            merakiapi.updateswitchport(api_key, switch_port["serial"], switch_port["number"], switch_port["name"],
                                       switch_port["tags"], switch_port["enabled"], switch_port["type"],
                                       switch_port["vlan"], switch_port["voiceVlan"], switch_port["allowedVlans"],
                                       switch_port["poeEnabled"], "", switch_port["rstpEnabled"], switch_port["stpGuard"],
                                       "")

            # progress_percent = '{:.1%}'.format(progress_count / progress_total)
            progress_count += 1
            print(progress_count)
            progress_percent = progress_count / progress_total * 100
            # print(progress_percent)
            yield "data:" + str(progress_percent) + "\n\n"
    return Response(generate(), mimetype ='text/event-stream')



    archive_path = os.path.abspath(os.path.join('app', 'archive'))
    shutil.copy(temp_path, archive_path)
    file_rename()
    archive_limit()

    #return "IT WORKS!"


def stream_template(template_name, **context):
    app.update_template_context(context)
    t = app.jinja_env.get_template(template_name)
    rv = t.stream(context)
    rv.disable_buffering()
    return rv



def generate():
    # configure()
    for progress in range(1):
        yield (progress_percent)
        sleep(1)




@configure_blueprint.route('/stream')
def stream_view():
    rows = generate()
    return Response(stream_template('step3.html', rows=rows))



if __name__ == "__main__":
    main()
